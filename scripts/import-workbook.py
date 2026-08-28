#!/usr/bin/env python3
"""
The spreadsheet, into the app. One-way, once.

`Výdaje 2026.xlsx` is the workbook Prosper replaced — eight months of rows, one
column pair per bucket, `popis` beside every amount. This reads it and writes a
**new backup**: the one exported from Settings, with the workbook's history
added to it. Load the result through Nastavení → Data → Načíst zálohu.

    pip install openpyxl
    python scripts/import-workbook.py \
        "Výdaje 2026.xlsx" prosper-zaloha.json prosper-zaloha-s-excelem.json \
        --skip-month 2026-08

Four things it will not do, and each one is the point:

  * **It never edits a row that is already in the backup.** Every existing row
    is copied through byte-identical, so `importBackup`'s last-write-wins merge
    finds `updatedAt` equal and `deviceId` equal, decides the incoming row does
    not win, and skips it. The import is additive by construction.
  * **It never lands on a month the app already has rows in.** That is asserted,
    not de-duplicated: pass `--skip-month` for the month being typed by hand and
    the two sets cannot meet. Without it the script refuses to write anything.
  * **It never invents a day.** The workbook has no dates — only which month
    sheet a row sat on, which is the observation that made tracking impossible
    in Excel — so every row is dated the **1st of its month** and its note says
    so. Spreading them across the month would look better and be fiction.
  * **It never guesses `isOneOff`.** That flag decides what "a normal month
    costs", and a guess would corrupt the one figure the Trimming law reads.

Money is integer haléře throughout — `Decimal(str(v)) * 100`, never float — and
every column is checked against the total the sheet computes for itself before
a single row is written.
"""

from __future__ import annotations

import argparse
import json
import secrets
import sys
from decimal import Decimal
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - a message beats a traceback
    sys.exit("openpyxl is missing. `pip install openpyxl`, then run this again.")

# Czech month sheets, in order. The workbook names its sheets and nothing else
# says which month a row belongs to.
SHEETS = [
    "LEDEN",
    "ÚNOR",
    "BŘEZEN",
    "DUBEN",
    "KVĚTEN",
    "ČERVEN",
    "ČERVENEC",
    "SRPEN",
    "ZÁŘÍ",
    "ŘÍJEN",
    "LISTOPAD",
    "PROSINEC",
]

# (amount column, bucket name). The `popis` for a bucket is the column after
# its amount. Q and R are CELKEM and VÝDAJE — formulas, not rows — and U/V is
# "Dočasně zrušeno", standing payments that were cancelled. None is read.
PAIRS = [
    (1, "PŘÍJEM"),
    (3, "DARY"),
    (5, "JÍDLO"),
    (7, "BYDLENÍ"),
    (9, "INVESTICE DO MĚ"),
    (11, "LIFESTYLE"),
    (13, "PROJEKTY"),
    (15, "OSTATNÍ"),
]

CELKEM_COLUMN = 17
HEADER_ROW = 1
TOTALS_ROW = 2
FIRST_DATA_ROW = 3

NOTE = "Import z „Výdaje 2026.xlsx“ — den v měsíci není v sešitu zapsán"

# The one bucket whose rows are money arriving. Everything else is money out —
# and a negative figure in a spending column is a refund, which Prosper already
# reads as an inflow filed under the bucket it came back to.
INCOME_BUCKET = "PŘÍJEM"


# ── UUIDv7, the same shape as domain/ids.ts ─────────────────────────────────
# 48-bit big-endian millisecond timestamp, version 7, a 12-bit counter, the
# variant bits, random for the rest. Ids sort by creation time, which is what
# the tape uses to order two rows written on the same day.

_sequence = 0


def uuidv7(ms: int) -> str:
    global _sequence
    _sequence = (_sequence + 1) & 0xFFF
    raw = bytearray(16)
    raw[0:6] = ms.to_bytes(6, "big")
    raw[6] = 0x70 | ((_sequence >> 8) & 0x0F)
    raw[7] = _sequence & 0xFF
    random = secrets.token_bytes(8)
    raw[8] = 0x80 | (random[0] & 0x3F)
    raw[9:16] = random[1:8]
    hexed = raw.hex()
    return f"{hexed[0:8]}-{hexed[8:12]}-{hexed[12:16]}-{hexed[16:20]}-{hexed[20:32]}"


def minor(value) -> int:
    """Koruny as they sit in the sheet -> integer haléře. No float arithmetic."""
    return int((Decimal(str(value)) * 100).to_integral_value())


def read_month(worksheet, month: str) -> list[tuple[str, str, object, str]]:
    """Every (month, bucket, koruny, popis) on one sheet."""
    rows = []
    for column, bucket in PAIRS:
        for r in range(FIRST_DATA_ROW, worksheet.max_row + 1):
            amount = worksheet.cell(row=r, column=column).value
            popis = worksheet.cell(row=r, column=column + 1).value
            if amount is None:
                if popis is not None:
                    raise SystemExit(
                        f"{worksheet.title}!{r}: a description with no amount: {popis!r}"
                    )
                continue
            if not isinstance(amount, (int, float)):
                raise SystemExit(f"{worksheet.title}!{r}: {amount!r} is not a number")
            rows.append((month, bucket, amount, str(popis or "").strip()))
    return rows


def check_against_the_sheet(worksheet, month: str, rows) -> None:
    """Each column against the total the sheet computes for itself."""
    for column, bucket in PAIRS:
        stated = worksheet.cell(row=TOTALS_ROW, column=column).value
        if stated is None:
            continue
        parsed = sum(Decimal(str(a)) for m, b, a, _ in rows if m == month and b == bucket)
        if Decimal(str(stated)) != parsed:
            raise SystemExit(
                f"{worksheet.title} {bucket}: parsed {parsed}, the sheet says {stated}. "
                "Nothing was written."
            )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    parser.add_argument("workbook", type=Path, help="Výdaje 2026.xlsx")
    parser.add_argument("backup", type=Path, help="a backup exported from Nastavení → Data")
    parser.add_argument("out", type=Path, help="where to write the merged backup")
    parser.add_argument("--year", type=int, default=2026, help="the year the sheets are (2026)")
    parser.add_argument(
        "--skip-month",
        action="append",
        default=[],
        metavar="YYYY-MM",
        help="a month to leave out — the current one, if it is being typed by hand",
    )
    args = parser.parse_args()

    backup = json.loads(args.backup.read_text(encoding="utf-8"))
    if backup.get("format") != "finance-backup":
        raise SystemExit(f"{args.backup} is not a Prosper backup.")

    # The live account and the live buckets. A backup can hold deleted twins
    # from an earlier seed, and pointing a row at one of those would file the
    # history into a category that no longer exists.
    accounts = [a for a in backup["accounts"] if not a["isDeleted"]]
    if len(accounts) != 1:
        raise SystemExit(f"expected one live account in the backup, found {len(accounts)}")
    account = accounts[0]

    categories = {
        c["name"]: c for c in backup["categories"] if not c["isDeleted"] and not c["isArchived"]
    }
    for _, bucket in PAIRS:
        if bucket not in categories:
            raise SystemExit(f"the backup has no live category named {bucket!r}")

    workbook = openpyxl.load_workbook(args.workbook, data_only=True)

    months = {
        sheet: f"{args.year}-{index + 1:02d}"
        for index, sheet in enumerate(SHEETS)
        if sheet in workbook.sheetnames and f"{args.year}-{index + 1:02d}" not in args.skip_month
    }
    if not months:
        raise SystemExit("every sheet was skipped — nothing to import.")

    # Not a de-duplication pass: a guarantee that one is not needed. If the app
    # already holds a row in a month being imported, importing would count that
    # month twice, and the answer is `--skip-month` rather than a fuzzy match.
    occupied = {t["date"][:7] for t in backup["txns"] if not t["isDeleted"]}
    clash = sorted(occupied & set(months.values()))
    if clash:
        raise SystemExit(
            f"the app already has transactions in {', '.join(clash)}. "
            f"Re-run with {' '.join(f'--skip-month {m}' for m in clash)}."
        )

    rows: list[tuple[str, str, object, str]] = []
    for sheet, month in months.items():
        rows += read_month(workbook[sheet], month)
    for sheet, month in months.items():
        check_against_the_sheet(workbook[sheet], month, rows)

    # ── the rows ─────────────────────────────────────────────────────────────
    stamp = 1_756_400_000_000  # a fixed base, so ids keep the workbook's order
    written_at = backup.get("exportedAt", "2026-01-01T00:00:00.000Z")
    device_id = account["deviceId"]

    txns = []
    for index, (month, bucket, koruny, popis) in enumerate(rows, start=1):
        magnitude = minor(koruny)
        txns.append(
            {
                "id": uuidv7(stamp + index),
                "accountId": account["id"],
                "date": f"{month}-01",
                "amount": magnitude if bucket == INCOME_BUCKET else -magnitude,
                "categoryId": categories[bucket]["id"],
                "payee": popis,
                "note": NOTE,
                "transferPairId": None,
                "source": "manual",
                "isCleared": False,
                "isOneOff": False,
                "owedAmount": None,
                "owedBy": None,
                "settledByTxnId": None,
                "scheduleId": None,
                "createdAt": written_at,
                "updatedAt": written_at,
                "deviceId": device_id,
                "isDeleted": False,
            }
        )

    # ── merge, additively ────────────────────────────────────────────────────
    original = json.dumps(backup["txns"], sort_keys=True)
    merged = dict(backup)
    merged["txns"] = backup["txns"] + txns
    assert json.dumps(merged["txns"][: len(backup["txns"])], sort_keys=True) == original
    assert len({t["id"] for t in merged["txns"]}) == len(merged["txns"]), "id collision"
    for key in ("accounts", "categories", "goals", "monthTargets", "holdings",
                "valuations", "schedules", "reconciliations", "dayMarks"):
        assert merged[key] is backup[key], f"{key} was rewritten"

    args.out.write_text(json.dumps(merged, ensure_ascii=False, indent="\t"), encoding="utf-8")

    # ── the report ───────────────────────────────────────────────────────────
    skipped = [s for s in SHEETS if s in workbook.sheetnames and s not in months]
    print(f"read      {len(rows)} rows from {args.workbook.name} "
          f"({min(months.values())} … {max(months.values())})")
    if skipped:
        print(f"left out  {', '.join(skipped)}")
    print(f"imported  {len(txns)} transactions")
    print()
    print(f"  {'month':<9} {'income':>12} {'outflow':>12} {'net':>12} {'sheet CELKEM':>14}")
    for sheet, month in months.items():
        month_rows = [t for t in txns if t["date"].startswith(month)]
        income = sum(t["amount"] for t in month_rows if t["amount"] > 0)
        outflow = sum(t["amount"] for t in month_rows if t["amount"] < 0)
        stated = workbook[sheet].cell(row=TOTALS_ROW, column=CELKEM_COLUMN).value
        net = income + outflow
        agrees = stated is None or net == minor(stated)
        print(f"  {month:<9} {income / 100:>12,.0f} {outflow / 100:>12,.0f} {net / 100:>12,.0f}"
              f" {(stated or 0):>14,.0f}   {'ok' if agrees else 'MISMATCH'}")

    imported_net = sum(t["amount"] for t in txns)
    existing_net = sum(t["amount"] for t in backup["txns"] if not t["isDeleted"])
    opening = account["openingBalance"]
    print()
    print(f"  balance before   {(opening + existing_net) / 100:>14,.2f} Kč")
    print(f"  imported         {imported_net / 100:>14,.2f} Kč")
    print(f"  balance after    {(opening + existing_net + imported_net) / 100:>14,.2f} Kč")
    print()
    print(f"written: {args.out}")
    print("The account's opening balance still says what it said. Eight months of history")
    print("hanging off a zero is not a balance — set it in Nastavení → Účet.")


if __name__ == "__main__":
    main()
